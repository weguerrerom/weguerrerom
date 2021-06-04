import os
#import Pillow 
from openpyxl import load_workbook
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties



path = 'G:\\Mi unidad\\WSP Ingenieria\\P1378\\2021\\Ajuste Informe suelos\\A. GEOLOGÍA\\A.4 FCH_PRCS_MORFODINAMICOS\\A.4.2 FORMATOS EXCEL\\OneDrive_4_29-4-2021\\pRUEBA\\' #Finalizar con doble \

for archivo in os.listdir(path):
    if archivo[-5:].lower() == ".xlsx":
        Libro = load_workbook(str(path + archivo))
        hojas = Libro.sheetnames
        for hoja in hojas:
            Hoja_act = Libro[str(hoja)]
            Hoja_act.print_options.horizontalCentered = True
            Hoja_act.print_options.verticalCentered = True
            Hoja_act.sheet_properties.fitToWidth = 1
            Hoja_act.oddFooter.right.text = "&[Page] of &[Pages]"
            Hoja_act.oddFooter.center.text = "&[File]"
        Libro.save(str(path + archivo+"1.xlsx"))
