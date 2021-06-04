import os
# from openpyxl import load_workbook
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from editpyxl import Workbook


wb = Workbook()

path = 'G:\\Mi unidad\\WSP Ingenieria\\P1378\\2021\\Archivos a PDF\\wetransfer-b94340\\h.natural\\Modificados Python\\' #Finalizar con doble \

for archivo in os.listdir(path):
    if archivo[-5:].lower() == ".xlsx":
        Path_file = path + str(archivo)
        Libro = wb.open(
            Path_file)
        # Libro = load_workbook(str(path + archivo))
        hojas = Libro.sheetnames
        for indice, hoja in enumerate(hojas, start = 1):
            if indice == 3:
                break
            Hoja_act = Libro[str(hoja)]
            # Hoja_act.print_options.horizontalCentered = True
            # Hoja_act.print_options.verticalCentered = True
            Hoja_act.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage= True, autoPageBreaks = False)
            # Hoja_act.oddFooter.right.text = "&[Page] of &[Pages]"
            # Hoja_act.oddFooter.center.text = "&[File]"
        Libro.save(str(path + archivo+"1.xlsx"))
